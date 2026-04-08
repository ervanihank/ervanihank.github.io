#!/usr/bin/env python3
"""
Generate book data from the Goodreads CSV export.
Reads the imported Goodreads CSV and outputs generated book data for the journal.

Usage:
  python3 generate_book_dataset.py

The script:
1. Reads goodreads_library_export.csv
2. Filters for books with status="read" (from Exclusive Shelf column)
3. Resolves author country from local map and internet lookup (Wikidata)
4. Resolves English titles from internet APIs and cached overrides
5. Adds cover image URLs from ISBN via OpenLibrary Covers
6. Generates book-data.js with all read books

Optional: Create book-country-map.json to override auto-detected countries:
{
  "Elena Ferrante": "italy",
  "Orhan Pamuk": "turkey"
}

Author: Script for enk_website journal
"""

import csv
import difflib
import html
import json
import re
import sys
import urllib.parse
import urllib.request
import unicodedata
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]


# Known authors and their countries (expand as needed)
AUTHOR_COUNTRY_MAP = {
    "Elena Ferrante": "italy",
    "Orhan Pamuk": "turkey",
    "Yaşar Kemal": "turkey",
    "Ahmet Hamdi Tanpınar": "turkey",
    "Nazan Bekiroğlu": "turkey",
    "Oğuz Atay": "turkey",
    "Vigdis Hjorth": "norway",
    "Leo Tolstoy": "russia",
    "Fyodor Dostoevsky": "russia",
    "Stefan Zweig": "austria",
    "Hermann Hesse": "germany",
    "William Shakespeare": "unitedKingdom",
    "John Berger": "unitedKingdom",
    "Oliver Sacks": "unitedKingdom",
    "Maggie O'Farrell": "unitedKingdom",
    "Magda Szabó": "hungary",
    "Ivan Goncharov": "russia",
    "Stendhal": "france",
    "F. Scott Fitzgerald": "unitedStates",
    "William Golding": "unitedKingdom",
    "Anthony Burgess": "unitedKingdom",
    "Mikhail Bulgakov": "russia",
    "Vladimir Nabokov": "russia",
    "Emily Brontë": "unitedKingdom",
    "Ursula K. Le Guin": "unitedStates",
    "Annie Ernaux": "france",
    "Johann Hari": "unitedKingdom",
    "Irvin D. Yalom": "unitedStates",
    "Viktor E. Frankl": "austria",
    "Byung-Chul Han": "germany",
    "Miguel Ruiz": "mexico",
    "Erlend Loe": "norway",
    "Michiko Aoyama": "japan",
    "Han Kang": "southKorea",
    "Paul Auster": "unitedStates",
    "Susan Sontag": "unitedStates",
    "José Eduardo Agualusa": "angola",
    "Theo Angelopoulos": "greece",
    "Pawel Pawlikowski": "poland",
    "Rob Reiner": "unitedStates",
    "Romain Gary": "france",
    "Pınar Kür": "turkey",
    "Osamu Dazai": "japan",
    "Dino Buzzati": "italy",
    "Amin Maalouf": "lebanon",
    "José Saramago": "portugal",
    "J.M. Coetzee": "southAfrica",
    "Ivo Andrić": "bosniaAndHerzegovina",
    "Şule Gürbüz": "turkey",
    "Kenzaburō Ōe": "japan",
    "Nermin Yıldırım": "turkey",
    "Şermin Yaşar": "turkey",
    "John Steinbeck": "unitedStates",
    "Abdulrazak Gurnah": "tanzania",
    "Wilhelm Schmid": "germany",
    "Dominique Bona": "france",
    "Alain de Botton": "switzerland",
    "Daniel Pennac": "france",
    "Ferhan Şensoy": "turkey",
    "Alejandro Zambra": "chile",
    "Nikos Kazantzakis": "greece",
    "Ian McEwan": "unitedKingdom",
    "Miguel de Cervantes Saavedra": "spain",
    "Buket Uzuner": "turkey",
    "Sandro Veronesi": "italy",
    "Brenda Lozano": "mexico",
}

COUNTRY_NAME_TO_CODE = {
    "Turkey": "turkey",
    "Turkiye": "turkey",
    "France": "france",
    "Japan": "japan",
    "United States": "unitedStates",
    "USA": "unitedStates",
    "United Kingdom": "unitedKingdom",
    "UK": "unitedKingdom",
    "Germany": "germany",
    "Poland": "poland",
    "Russia": "russia",
    "Greece": "greece",
    "Netherlands": "netherlands",
    "Belgium": "belgium",
    "Austria": "austria",
    "Spain": "spain",
    "Italy": "italy",
    "Sweden": "sweden",
    "Norway": "norway",
    "Denmark": "denmark",
    "Finland": "finland",
    "Hungary": "hungary",
    "Czech Republic": "czechRepublic",
    "Ireland": "ireland",
    "Canada": "canada",
    "Australia": "australia",
    "China": "china",
    "South Korea": "southKorea",
    "India": "india",
    "Argentina": "argentina",
    "Brazil": "brazil",
    "Mexico": "mexico",
    "Israel": "israel",
    "Egypt": "egypt",
    "Lebanon": "lebanon",
    "Portugal": "portugal",
    "Angola": "angola",
    "Colombia": "colombia",
    "Bosnia and Herzegovina": "bosniaAndHerzegovina",
    "Serbia": "serbia",
    "Chile": "chile",
    "Iran": "iran",
    "Romania": "romania",
    "South Africa": "southAfrica",
    "Tanzania": "tanzania",
    "Switzerland": "switzerland",
    "United States of America": "unitedStates",
    "U.S.A.": "unitedStates",
    "Russian Empire": "russia",
    "Soviet Union": "russia",
    "USSR": "russia",
    "Ottoman Empire": "turkey",
    "German Reich": "germany",
    "German Empire": "germany",
    "Saxe-Weimar-Eisenach": "germany",
    "United Kingdom of Great Britain and Ireland": "unitedKingdom",
}

COUNTRY_CODE_TO_LABEL = {
    "turkey": "Turkey",
    "france": "France",
    "japan": "Japan",
    "unitedStates": "United States",
    "unitedKingdom": "United Kingdom",
    "germany": "Germany",
    "poland": "Poland",
    "russia": "Russia",
    "greece": "Greece",
    "netherlands": "Netherlands",
    "belgium": "Belgium",
    "austria": "Austria",
    "spain": "Spain",
    "italy": "Italy",
    "sweden": "Sweden",
    "norway": "Norway",
    "denmark": "Denmark",
    "finland": "Finland",
    "hungary": "Hungary",
    "czechRepublic": "Czech Republic",
    "ireland": "Ireland",
    "canada": "Canada",
    "australia": "Australia",
    "china": "China",
    "southKorea": "South Korea",
    "india": "India",
    "argentina": "Argentina",
    "brazil": "Brazil",
    "mexico": "Mexico",
    "israel": "Israel",
    "egypt": "Egypt",
    "lebanon": "Lebanon",
    "portugal": "Portugal",
    "angola": "Angola",
    "colombia": "Colombia",
    "bosniaAndHerzegovina": "Bosnia and Herzegovina",
    "serbia": "Serbia",
    "chile": "Chile",
    "iran": "Iran",
    "romania": "Romania",
    "southAfrica": "South Africa",
    "tanzania": "Tanzania",
    "switzerland": "Switzerland",
}

WIKIDATA_API = "https://www.wikidata.org/w/api.php"
CACHE_DIR = ROOT / "data" / "cache"
COUNTRY_CACHE_FILE = CACHE_DIR / ".author-country-cache.json"
TITLE_CACHE_FILE = CACHE_DIR / ".book-title-cache.json"
COVER_CACHE_FILE = CACHE_DIR / ".book-cover-cache.json"
TITLE_OVERRIDE_FILE = ROOT / "data" / "config" / "book-title-map.json"
POSTS_XLSX_FILE = ROOT / "data" / "source" / "Posts.xlsx"
INSTAGRAM_POSTS_HTML_FILE = Path(
    "instagram-biblionih-2026-03-31-yhcuNyrx/your_instagram_activity/media/posts_1.html"
)
BOOK_COVERS_DIR = ROOT / "assets" / "images" / "covers" / "books"
BOOKCLUB_DATA_FILE = ROOT / "data" / "generated" / "auto" / "bookclub-data.js"

COUNTRY_COMPACT_ALIASES = {
    "usa": "unitedStates",
    "us": "unitedStates",
    "unitedstates": "unitedStates",
    "unitedstatesofamerica": "unitedStates",
    "uk": "unitedKingdom",
    "unitedkingdom": "unitedKingdom",
    "unitedkingdomofgreatbritainandireland": "unitedKingdom",
    "russianempire": "russia",
    "sovietunion": "russia",
    "ussr": "russia",
    "ottomanempire": "turkey",
    "germanreich": "germany",
    "germanempire": "germany",
    "saxeweimareisenach": "germany",
}

EN_NOTE_TR_ONLY = (
    "I only wrote my notes for this book in Turkish and preferred not to force an English translation. "
    "Please check the Turkish version."
)
TR_NOTE_EN_ONLY = (
    "Bu kitap icin notlarimi sadece Ingilizce yazdim ve zoraki bir Turkce ceviri yapmak istemedim. "
    "Lutfen Ingilizce versiyona bakiniz."
)
EN_QUOTE_TR_ONLY = "I only kept quotations for this book in Turkish. Please check the Turkish version."
TR_QUOTE_EN_ONLY = "Bu kitap icin alintilari sadece Ingilizce tuttum. Lutfen Ingilizce versiyona bakiniz."


def slugify(text):
    """Create a URL-safe slug from text."""
    return (
        text.lower()
        .replace(" ", "-")
        .replace("'", "")
        .replace("\"", "")
        .replace("é", "e")
        .replace("ş", "s")
        .replace("ı", "i")
        .replace("ğ", "g")
        .replace("ü", "u")
        .replace("ç", "c")
        .replace("ö", "o")
    )


def keywordify(text):
    """Normalize free text into a compact keyword token."""
    if not text:
        return None
    value = slugify(text)
    return value if value else None


def normalize_text(value):
    """Normalize strings for robust matching across data sources."""
    text = (value or "").strip().lower()
    replacements = {
        "ı": "i",
        "İ": "i",
        "ö": "o",
        "ü": "u",
        "ş": "s",
        "ç": "c",
        "ğ": "g",
        "â": "a",
        "î": "i",
        "û": "u",
    }
    for src, target in replacements.items():
        text = text.replace(src, target)
    text = html.unescape(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text


def compact_text(value):
    """Compact normalized text for fuzzy comparisons."""
    return "".join(ch for ch in normalize_text(value) if ch.isalnum())


def canonical_author_key(value):
    """Normalize author names while ignoring parenthetical aliases."""
    text = (value or "").strip()
    text = re.sub(r"\s*\([^)]*\)\s*$", "", text).strip()
    return compact_text(text)


def split_keywords(value):
    """Split keyword cells separated by commas, semicolons, pipes, or new lines."""
    raw = (value or "").strip()
    if not raw:
        return []
    parts = re.split(r"[,;|\n]+", raw)
    return [part.strip() for part in parts if part and part.strip()]


def split_quotes(value):
    """Extract quote candidates from a free-form quote field."""
    raw = html.unescape(str(value or ""))
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    raw = raw.replace("“", '"').replace("”", '"').replace("‘", "'").replace("’", "'")
    raw = re.sub(r"\u00ad", "", raw).strip()
    if not raw:
        return []

    def clean_fragment(fragment):
        text = fragment.strip()
        text = re.sub(r"^[-–—•\s]+", "", text)
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"\s+([,.;:!?])", r"\1", text)
        text = re.sub(r"([\(\[\{])\s+", r"\1", text)
        text = re.sub(r"\s+([\)\]\}])", r"\1", text)
        text = text.strip('"\'“”‘’` ')
        return text.strip()

    lines = [line.strip() for line in raw.split("\n") if line.strip()]
    if not lines:
        return []

    blocks = []
    current = []
    for line in lines:
        if current and re.match(r"^[-–—•]\s*", line):
            blocks.append(current)
            current = [line]
        else:
            current.append(line)
    if current:
        blocks.append(current)

    quotes = []
    for block in blocks:
        joined_parts = []
        for line in block:
            piece = re.sub(r"^[-–—•]\s*", "", line).strip()
            if not piece:
                continue
            if joined_parts and joined_parts[-1].endswith("-") and piece[:1].islower():
                joined_parts[-1] = joined_parts[-1][:-1] + piece
            else:
                joined_parts.append(piece)

        joined = " ".join(joined_parts)
        cleaned = clean_fragment(joined)
        if cleaned and cleaned not in quotes:
            quotes.append(cleaned)

    return quotes


def split_quotes_by_language(quotes):
    """Split quote list into EN/TR buckets using lightweight language detection."""
    en_quotes = []
    tr_quotes = []
    for quote in (quotes or []):
        cleaned = (quote or "").strip()
        if not cleaned:
            continue
        if is_likely_turkish_text(cleaned):
            tr_quotes.append(cleaned)
        else:
            en_quotes.append(cleaned)
    return en_quotes, tr_quotes


def parse_js_array_variable(file_path, variable_name):
    """Parse `const variableName = [...]` from a JS file."""
    if not file_path.exists():
        return []

    content = file_path.read_text(encoding="utf-8")
    marker = f"const {variable_name} = "
    start = content.find(marker)
    if start < 0:
        return []
    start += len(marker)
    end = content.rfind(";")
    if end <= start:
        return []

    try:
        parsed = json.loads(content[start:end])
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def load_bookclub_lookup():
    """Build a lookup for matching Goodreads books to book club note pages."""
    entries = parse_js_array_variable(BOOKCLUB_DATA_FILE, "bookClubEntries")
    lookup = {}

    for entry in entries:
        slug = (entry.get("slug") or "").strip()
        if not slug:
            continue

        keys = {
            compact_text(entry.get("book") or ""),
            compact_text(entry.get("englishTitle") or ""),
        }

        for key in keys:
            if not key:
                continue
            lookup.setdefault(key, []).append(entry)

    return lookup


def load_local_cover_lookup():
    """Build compact-title lookup for local book cover files."""
    lookup = {}
    if not BOOK_COVERS_DIR.exists():
        return lookup

    for path in BOOK_COVERS_DIR.iterdir():
        if path.suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp"}:
            continue
        key = compact_text(path.stem)
        if key:
            lookup[key] = f"assets/images/covers/books/{path.name}"

    return lookup


def resolve_local_cover(title_candidates, cover_lookup):
    """Find exact-matching local cover from the local covers folder."""
    if not cover_lookup:
        return None

    candidate_keys = [compact_text(candidate) for candidate in title_candidates if candidate]
    candidate_keys = [key for key in candidate_keys if key]
    if not candidate_keys:
        return None

    for key in candidate_keys:
        if key in cover_lookup:
            return cover_lookup[key]

    return None


def load_posts_excel_rows():
    """Load curated notes/quotes metadata from Posts.xlsx."""
    if not POSTS_XLSX_FILE.exists():
        return []

    try:
        workbook = load_workbook(POSTS_XLSX_FILE, read_only=True, data_only=True)
    except Exception as exc:
        print(f"Warning: Could not read {POSTS_XLSX_FILE}: {exc}")
        return []

    if not workbook.sheetnames:
        return []

    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_map = {str(value).strip().lower(): idx for idx, value in enumerate(headers) if value}

    def value_from(row, header_name):
        idx = header_map.get(header_name.lower())
        if idx is None or idx >= len(row):
            return ""
        cell_value = row[idx]
        if cell_value is None:
            return ""
        return str(cell_value).strip()

    def value_from_any(row, header_names):
        for header_name in header_names:
            value = value_from(row, header_name)
            if value:
                return value
        return ""

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        title_tr = value_from(row, "Turkish Name")
        title_en = value_from(row, "English Title")
        author = value_from(row, "Author")
        if not (title_tr or title_en):
            continue

        note = value_from(row, "Notes")
        quote = value_from(row, "Quotes")
        note_en = value_from_any(row, ["Notes EN", "English Notes", "Note EN"])
        note_tr = value_from_any(row, ["Notes TR", "Turkish Notes", "Note TR"])
        quote_en = value_from_any(row, ["Quotes EN", "English Quotes", "Quote EN"])
        quote_tr = value_from_any(row, ["Quotes TR", "Turkish Quotes", "Quote TR"])
        place_date = value_from(row, "place/date")
        add_to_bookclub = value_from(row, "Add also to book club")
        keywords = value_from(row, "keywords")

        records.append(
            {
                "title_tr": title_tr,
                "title_en": title_en,
                "author": author,
                "note": note,
                "note_en": note_en,
                "note_tr": note_tr,
                "quotes": split_quotes(quote),
                "quotes_en": split_quotes(quote_en),
                "quotes_tr": split_quotes(quote_tr),
                "place_date": place_date,
                "add_to_bookclub": add_to_bookclub.lower() in {"yes", "true", "1", "evet"},
                "keywords": split_keywords(keywords),
            }
        )

    return records


def build_posts_excel_lookup(records):
    """Index post metadata rows by normalized title and title+author for fast matching."""
    lookup = {}
    for record in records:
        keys = {
            compact_text(record.get("title_tr")),
            compact_text(record.get("title_en")),
        }
        author_key = canonical_author_key(record.get("author"))
        for key in keys:
            if not key:
                continue
            lookup.setdefault(key, []).append(record)
            if author_key:
                lookup.setdefault(f"{key}||{author_key}", []).append(record)
    return lookup


def find_posts_candidates(candidate_keys, posts_lookup):
    """Resolve best Posts.xlsx candidates from exact and fuzzy compact-title matching."""
    candidates = []
    seen = set()

    for key in candidate_keys:
        if not key:
            continue
        for item in posts_lookup.get(key, []):
            obj_id = id(item)
            if obj_id not in seen:
                seen.add(obj_id)
                candidates.append(item)

    if candidates:
        return candidates

    available_keys = list(posts_lookup.keys())
    for key in candidate_keys:
        if not key:
            continue
        close = difflib.get_close_matches(key, available_keys, n=3, cutoff=0.84)
        for matched_key in close:
            for item in posts_lookup.get(matched_key, []):
                obj_id = id(item)
                if obj_id not in seen:
                    seen.add(obj_id)
                    candidates.append(item)

    return candidates


def find_exact_posts_candidates(candidate_keys, author_key, posts_lookup):
    """Resolve only exact title+author Posts.xlsx matches."""
    if not author_key:
        return []

    candidates = []
    seen = set()
    for key in candidate_keys:
        if not key:
            continue
        exact_key = f"{key}||{author_key}"
        for item in posts_lookup.get(exact_key, []):
            obj_id = id(item)
            if obj_id not in seen:
                seen.add(obj_id)
                candidates.append(item)
    return candidates


def parse_instagram_posts_html():
    """Parse Instagram posts export for caption fallback text."""
    if not INSTAGRAM_POSTS_HTML_FILE.exists():
        return []

    try:
        content = INSTAGRAM_POSTS_HTML_FILE.read_text(encoding="utf-8")
    except Exception as exc:
        print(f"Warning: Could not read {INSTAGRAM_POSTS_HTML_FILE}: {exc}")
        return []

    chunks = content.split('<div class="pam _3-95 _2ph- _a6-g uiBoxWhite noborder">')
    posts = []
    for chunk in chunks[1:]:
        caption_match = re.search(r'<h2[^>]*>(.*?)</h2>', chunk, re.DOTALL)
        if not caption_match:
            continue

        raw_caption = caption_match.group(1)
        caption = re.sub(r"<[^>]+>", " ", raw_caption)
        caption = html.unescape(caption)
        caption = re.sub(r"\s+", " ", caption).strip()
        if not caption:
            continue

        images = re.findall(r'href="(media/posts/[^"]+)"', chunk)
        date_match = re.search(r'<div class="_3-94 _a6-o">(.*?)</div>', chunk)
        date_label = html.unescape(date_match.group(1)).strip() if date_match else ""

        posts.append(
            {
                "caption": caption,
                "caption_compact": compact_text(caption),
                "images": images,
                "date": date_label,
            }
        )

    return posts


def find_instagram_post_match(title_tr, title_en, author, instagram_posts):
    """Find best matching Instagram post caption for a book."""
    if not instagram_posts:
        return None

    title_keys = [compact_text(title_tr), compact_text(title_en)]
    title_keys = [key for key in title_keys if key and len(key) >= 5]
    author_key = canonical_author_key(author)
    if not title_keys or not author_key:
        return None

    for post in instagram_posts:
        caption_compact = post.get("caption_compact", "")
        if not caption_compact:
            continue

        title_hit = any(key in caption_compact for key in title_keys)
        author_hit = bool(author_key and author_key in caption_compact)
        if title_hit and author_hit:
            return post

    return None


def resolve_post_metadata(title_tr, title_en, author, posts_lookup, instagram_posts):
    """Resolve optional metadata from Posts.xlsx and Instagram HTML export.

    Curated notes and quotes only attach when both book title and author match.
    """
    candidate_keys = [compact_text(title_tr), compact_text(title_en)]
    candidate_keys = [key for key in candidate_keys if key]
    author_key = canonical_author_key(author)

    candidates = find_exact_posts_candidates(candidate_keys, author_key, posts_lookup)

    selected = None
    if candidates:
        selected = candidates[0]

    instagram_match = find_instagram_post_match(title_tr, title_en, author, instagram_posts)
    return selected, instagram_match


def resolve_bookclub_url(title_tr, title_en, author, enabled, bookclub_lookup):
    """Create optional link to matching book club notes entry."""
    if not enabled:
        return None

    candidate_keys = [compact_text(title_tr), compact_text(title_en)]
    candidate_keys = [key for key in candidate_keys if key]
    author_key = canonical_author_key(author)

    for key in candidate_keys:
        matches = bookclub_lookup.get(key, [])
        if not matches:
            continue

        for match in matches:
            authors = [compact_text(item) for item in (match.get("authors") or []) if item]
            if authors and author_key and not any(author_key in item or item in author_key for item in authors):
                continue
            slug = (match.get("slug") or "").strip()
            if slug:
                return f"bookclub-notes.html?entry={urllib.parse.quote(slug)}"

    return None


def clean_isbn(raw):
    """Normalize Goodreads ISBN values like ="978...." into plain ISBN."""
    if not raw:
        return ""
    cleaned = raw.replace("=", "").replace('"', "").replace("'", "").strip()
    return "".join(ch for ch in cleaned if ch.isdigit() or ch in {"X", "x"}).upper()


def title_cache_key(book_id_source, author, title):
    """Stable cache key for title lookup results."""
    if book_id_source and book_id_source.isdigit():
        return f"book:{book_id_source}"
    return f"title:{author}|{title}"


def contains_turkish_chars(text):
    """Heuristic to detect Turkish-localized title strings."""
    if not text:
        return False
    turkish_chars = set("çğıöşüÇĞİÖŞÜ")
    return any(ch in turkish_chars for ch in text)


def is_likely_turkish_text(text):
    value = (text or "").strip()
    if not value:
        return False
    if contains_turkish_chars(value):
        return True
    normalized = normalize_text(value)
    markers = [" ve ", " bir ", " icin ", " bu ", " ile ", " gibi ", " daha "]
    scoped = f" {normalized} "
    return any(marker in scoped for marker in markers)


def build_bilingual_note_pair(note_en, note_tr):
    en = (note_en or "").strip()
    tr = (note_tr or "").strip()

    if en and tr:
        return en, tr
    if tr and not en:
        return EN_NOTE_TR_ONLY, tr
    if en and not tr:
        return en, TR_NOTE_EN_ONLY
    return "", ""


def build_bilingual_quote_pair(quotes_en, quotes_tr):
    en_list = [q for q in (quotes_en or []) if (q or "").strip()]
    tr_list = [q for q in (quotes_tr or []) if (q or "").strip()]

    if en_list and tr_list:
        return en_list, tr_list
    if tr_list and not en_list:
        return [EN_QUOTE_TR_ONLY], tr_list
    if en_list and not tr_list:
        return en_list, [TR_QUOTE_EN_ONLY]
    return [], []


def extract_parenthetical_title(title):
    """Extract title in parentheses if present, e.g. Local Title (Original Title)."""
    if "(" not in title or ")" not in title:
        return None
    start = title.rfind("(")
    end = title.rfind(")")
    if start >= 0 and end > start:
        inner = title[start + 1 : end].strip()
        return inner if inner else None
    return None


OPENLIBRARY_HAS_COVER_CACHE = {}


def openlibrary_has_cover(isbn):
    """Check whether OpenLibrary has a real cover for an ISBN."""
    if not isbn:
        return False

    if isbn in OPENLIBRARY_HAS_COVER_CACHE:
        return OPENLIBRARY_HAS_COVER_CACHE[isbn]

    try:
        params = {
            "bibkeys": f"ISBN:{isbn}",
            "format": "json",
            "jscmd": "data",
        }
        url = f"https://openlibrary.org/api/books?{urllib.parse.urlencode(params)}"
        payload = fetch_json(url)
        book_data = payload.get(f"ISBN:{isbn}", {}) if isinstance(payload, dict) else {}
        has_cover = bool(book_data.get("cover"))
    except Exception:
        has_cover = False

    OPENLIBRARY_HAS_COVER_CACHE[isbn] = has_cover
    return has_cover


def lookup_google_books_cover(isbn13, isbn, title, author):
    """Find likely cover URL via Google Books using ISBN first, then title/author."""
    title_candidates = []
    if title:
        title_candidates.append(title)

        parenthetical_title = extract_parenthetical_title(title)
        if parenthetical_title:
            title_candidates.append(parenthetical_title)

        # Remove trailing parenthetical parts like "(Series #1)".
        stripped_title = re.sub(r"\s*\([^)]*\)\s*$", "", title).strip()
        if stripped_title:
            title_candidates.append(stripped_title)

    seen_titles = set()
    unique_title_candidates = []
    for candidate in title_candidates:
        key = candidate.lower()
        if key in seen_titles:
            continue
        seen_titles.add(key)
        unique_title_candidates.append(candidate)

    queries = []
    if isbn13:
        queries.append(f"isbn:{isbn13}")
    if isbn and isbn != isbn13:
        queries.append(f"isbn:{isbn}")

    for candidate_title in unique_title_candidates:
        if author:
            queries.append(f"intitle:{candidate_title} inauthor:{author}")
        queries.append(f"intitle:{candidate_title}")

    if author:
        queries.append(f"inauthor:{author}")

    for query in queries:
        try:
            params = {
                "q": query,
                "maxResults": "5",
                "printType": "books",
            }
            url = f"https://www.googleapis.com/books/v1/volumes?{urllib.parse.urlencode(params)}"
            payload = fetch_json(url)
            items = payload.get("items", []) if isinstance(payload, dict) else []
        except Exception:
            continue

        for item in items:
            image_links = item.get("volumeInfo", {}).get("imageLinks", {})
            candidate = (
                image_links.get("thumbnail")
                or image_links.get("smallThumbnail")
                or image_links.get("small")
                or image_links.get("medium")
                or image_links.get("large")
            )
            if candidate:
                return candidate.replace("http://", "https://")

    return None


def lookup_openlibrary_search_cover(title, author):
    """Find cover via OpenLibrary search when ISBN is missing."""
    if not title:
        return None

    queries = []
    if author:
        queries.append({"title": title, "author": author, "limit": "5"})
    queries.append({"title": title, "limit": "5"})

    parenthetical_title = extract_parenthetical_title(title)
    if parenthetical_title:
        if author:
            queries.append({"title": parenthetical_title, "author": author, "limit": "5"})
        queries.append({"title": parenthetical_title, "limit": "5"})

    stripped_title = re.sub(r"\s*\([^)]*\)\s*$", "", title).strip()
    if stripped_title and stripped_title.lower() != title.lower():
        if author:
            queries.append({"title": stripped_title, "author": author, "limit": "5"})
        queries.append({"title": stripped_title, "limit": "5"})

    seen = set()
    for params in queries:
        key = tuple(sorted(params.items()))
        if key in seen:
            continue
        seen.add(key)

        try:
            url = f"https://openlibrary.org/search.json?{urllib.parse.urlencode(params)}"
            payload = fetch_json(url)
        except Exception:
            continue

        docs = payload.get("docs", []) if isinstance(payload, dict) else []
        for doc in docs:
            cover_id = doc.get("cover_i")
            if cover_id:
                return f"https://covers.openlibrary.org/b/id/{cover_id}-L.jpg"

    return None


def lookup_goodreads_cover(book_id):
    """Fetch cover URL from Goodreads book page metadata when available."""
    if not book_id or not str(book_id).strip().isdigit():
        return None

    url = f"https://www.goodreads.com/book/show/{str(book_id).strip()}"
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9",
        },
    )

    try:
        with urllib.request.urlopen(request, timeout=6) as response:
            page = response.read().decode("utf-8", errors="replace")
    except Exception:
        return None

    match = re.search(r'<meta\s+property="og:image"\s+content="([^"]+)"', page, re.I)
    if not match:
        return None

    candidate = html.unescape(match.group(1)).strip()
    if not candidate:
        return None

    return candidate.replace("http://", "https://")


def build_cover_url(row, cover_cache, cover_lookup_budget):
    """Build cover URL with stronger fallbacks for missing covers."""
    title = (row.get("Title") or "").strip()
    author = (row.get("Author") or "").strip()
    book_id = (row.get("Book Id") or "").strip()

    cache_key = f"book:{book_id}" if book_id.isdigit() else f"title:{author}|{title}"
    if cache_key in cover_cache and not book_id.isdigit():
        return cover_cache[cache_key] or None

    isbn13 = clean_isbn(row.get("ISBN13", ""))
    isbn = clean_isbn(row.get("ISBN", ""))
    best_isbn = isbn13 or isbn

    goodreads_cover = lookup_goodreads_cover(row.get("Book Id", ""))
    if goodreads_cover:
        cover_cache[cache_key] = goodreads_cover
        return goodreads_cover

    if cache_key in cover_cache:
        return cover_cache[cache_key] or None

    if best_isbn and openlibrary_has_cover(best_isbn):
        result = f"https://covers.openlibrary.org/b/isbn/{best_isbn}-L.jpg?default=false"
        cover_cache[cache_key] = result
        return result

    if cover_lookup_budget["remaining"] > 0:
        cover_lookup_budget["remaining"] -= 1

        google_cover = lookup_google_books_cover(isbn13, isbn, title, author)
        if google_cover:
            cover_cache[cache_key] = google_cover
            return google_cover

        openlibrary_cover = lookup_openlibrary_search_cover(title, author)
        if openlibrary_cover:
            cover_cache[cache_key] = openlibrary_cover
            return openlibrary_cover

    cover_cache[cache_key] = None
    return None


def load_custom_country_map():
    """Load optional book-country-map.json for manual overrides."""
    map_file = Path("book-country-map.json")
    if map_file.exists():
        try:
            with open(map_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: Could not load book-country-map.json: {e}")
            return {}
    return {}


def load_title_override_map():
    """Load optional manual English title overrides from book-title-map.json.

    Accepted keys:
    - Goodreads book id as string (recommended): "217898868": "What You Are Looking For Is in the Library"
    - Full original title fallback: "Aradığın Şey Kütüphanede Saklı": "What You Are Looking For Is in the Library"
    """
    if TITLE_OVERRIDE_FILE.exists():
        try:
            with open(TITLE_OVERRIDE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception as e:
            print(f"Warning: Could not load {TITLE_OVERRIDE_FILE}: {e}")
            return {}
    return {}


def load_country_cache():
    """Load cached internet lookups for author country."""
    if COUNTRY_CACHE_FILE.exists():
        try:
            with open(COUNTRY_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def load_title_cache():
    """Load cached internet lookups for English titles."""
    if TITLE_CACHE_FILE.exists():
        try:
            with open(TITLE_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def load_cover_cache():
    """Load cached internet lookups for cover URLs."""
    if COVER_CACHE_FILE.exists():
        try:
            with open(COVER_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def save_country_cache(cache):
    """Persist author-country lookup cache."""
    try:
        COUNTRY_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(COUNTRY_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"Warning: Could not write {COUNTRY_CACHE_FILE}: {exc}")


def save_title_cache(cache):
    """Persist English title lookup cache."""
    try:
        TITLE_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(TITLE_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"Warning: Could not write {TITLE_CACHE_FILE}: {exc}")


def save_cover_cache(cache):
    """Persist cover lookup cache."""
    try:
        COVER_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(COVER_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"Warning: Could not write {COVER_CACHE_FILE}: {exc}")


def fetch_json(url):
    """Fetch JSON with a basic User-Agent and timeout."""
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "enk-book-dataset-generator/1.0"},
    )
    with urllib.request.urlopen(request, timeout=4) as response:
        return json.loads(response.read().decode("utf-8"))


def lookup_google_books_title(isbn13, isbn, title, author):
    """Find likely English title via Google Books API."""
    queries = []
    if isbn13:
        queries.append(f"isbn:{isbn13}")
    if isbn and isbn != isbn13:
        queries.append(f"isbn:{isbn}")

    # Only use broad title query when ISBN is missing.
    if not queries:
        queries.append(f"intitle:{title} inauthor:{author}")

    for query in queries:
        params = {
            "q": query,
            "langRestrict": "en",
            "maxResults": "5",
            "printType": "books",
        }
        url = f"https://www.googleapis.com/books/v1/volumes?{urllib.parse.urlencode(params)}"

        try:
            payload = fetch_json(url)
        except Exception:
            continue

        items = payload.get("items", [])
        for item in items:
            info = item.get("volumeInfo", {})
            candidate = (info.get("title") or "").strip()
            language = (info.get("language") or "").strip().lower()
            if not candidate:
                continue
            if language.startswith("en"):
                return candidate

    return None


def lookup_goodreads_original_title(book_id):
    """Try to extract original/English title from Goodreads page metadata."""
    raw_id = (book_id or "").strip()
    if not raw_id.isdigit():
        return None

    url = f"https://www.goodreads.com/book/show/{raw_id}"
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9",
        },
    )

    try:
        with urllib.request.urlopen(request, timeout=6) as response:
            page = response.read().decode("utf-8", errors="replace")
    except Exception:
        return None

    patterns = [
        r'"originalTitle"\s*:\s*"([^"]+)"',
        r'Original\s+Title\s*</[^>]+>\s*<[^>]+>\s*([^<]+?)\s*</',
    ]

    for pattern in patterns:
        match = re.search(pattern, page, re.I | re.S)
        if not match:
            continue
        candidate = html.unescape(match.group(1)).strip()
        candidate = re.sub(r"\s+", " ", candidate)
        if candidate and not contains_turkish_chars(candidate):
            return candidate

    return None


def translate_turkish_title_to_english(title):
    """Translate unresolved Turkish title to English as a last-resort display fallback."""
    source = (title or "").strip()
    if not source or not contains_turkish_chars(source):
        return None

    params = {
        "client": "gtx",
        "sl": "tr",
        "tl": "en",
        "dt": "t",
        "q": source,
    }
    url = f"https://translate.googleapis.com/translate_a/single?{urllib.parse.urlencode(params)}"

    try:
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        with urllib.request.urlopen(request, timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception:
        return None

    if not isinstance(payload, list) or not payload:
        return None

    segments = payload[0] if isinstance(payload[0], list) else []
    translated = "".join(seg[0] for seg in segments if isinstance(seg, list) and seg and isinstance(seg[0], str)).strip()
    if not translated:
        return None

    # Keep fallback only when it actually differs from source title.
    if compact_text(translated) == compact_text(source):
        return None

    return translated


def translate_title_to_english_auto(title):
    """Translate title to English when source language is not English."""
    source = (title or "").strip()
    if not source:
        return None

    params = {
        "client": "gtx",
        "sl": "auto",
        "tl": "en",
        "dt": "t",
        "q": source,
    }
    url = f"https://translate.googleapis.com/translate_a/single?{urllib.parse.urlencode(params)}"

    try:
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        with urllib.request.urlopen(request, timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception:
        return None

    if not isinstance(payload, list) or not payload:
        return None

    segments = payload[0] if isinstance(payload[0], list) else []
    translated = "".join(seg[0] for seg in segments if isinstance(seg, list) and seg and isinstance(seg[0], str)).strip()
    if not translated:
        return None

    detected_lang = ""
    if len(payload) > 2 and isinstance(payload[2], str):
        detected_lang = payload[2].lower()

    if detected_lang.startswith("en"):
        return None

    if compact_text(translated) == compact_text(source):
        return None

    return translated


def resolve_english_title(row, title, author, title_overrides, title_cache, lookup_budget):
    """Resolve English title from override, cache, APIs, and heuristics."""
    book_id_source = (row.get("Book Id") or "").strip()
    key = title_cache_key(book_id_source, author, title)

    if book_id_source in title_overrides:
        return title_overrides[book_id_source]
    if title in title_overrides:
        return title_overrides[title]

    cached_title = title_cache.get(key)
    # Refresh stale cache entries where EN title is still identical to source title.
    if cached_title:
        if cached_title != title:
            return cached_title

    # Keep parenthetical extraction as a high-confidence local hint.
    parenthetical = extract_parenthetical_title(title)

    # Prevent very long runs by capping online lookups per generation.
    if lookup_budget["remaining"] <= 0:
        title_cache[key] = title
        return title

    isbn13 = clean_isbn(row.get("ISBN13", ""))
    isbn = clean_isbn(row.get("ISBN", ""))
    lookup_budget["remaining"] -= 1

    api_title = lookup_google_books_title(isbn13, isbn, title, author)

    if api_title:
        title_cache[key] = api_title
        return api_title

    goodreads_original = lookup_goodreads_original_title(book_id_source)
    if goodreads_original:
        title_cache[key] = goodreads_original
        return goodreads_original

    if parenthetical and not contains_turkish_chars(parenthetical):
        title_cache[key] = parenthetical
        return parenthetical

    title_cache[key] = title
    return title


def normalize_country_code(country_name):
    """Map country labels to journal country code."""
    if not country_name:
        return None

    # Already-normalized internal code.
    if country_name in COUNTRY_CODE_TO_LABEL:
        return country_name

    if country_name in COUNTRY_NAME_TO_CODE:
        return COUNTRY_NAME_TO_CODE[country_name]

    compact = "".join(ch for ch in country_name.lower() if ch.isalnum())
    if compact in COUNTRY_COMPACT_ALIASES:
        return COUNTRY_COMPACT_ALIASES[compact]

    for name, code in COUNTRY_NAME_TO_CODE.items():
        if name.lower() == country_name.lower():
            return code

    return compact if compact else None


def country_label_from_code(country_code):
    """Return display label from internal code."""
    if country_code in COUNTRY_CODE_TO_LABEL:
        return COUNTRY_CODE_TO_LABEL[country_code]
    return country_code


def lookup_country_via_wikidata(author):
    """Resolve author nationality country via Wikidata APIs."""
    search_params = {
        "action": "wbsearchentities",
        "format": "json",
        "language": "en",
        "type": "item",
        "limit": "5",
        "search": author,
    }
    search_url = f"{WIKIDATA_API}?{urllib.parse.urlencode(search_params)}"

    try:
        search_payload = fetch_json(search_url)
    except Exception:
        return None, None

    candidates = search_payload.get("search", [])
    if not candidates:
        return None, None

    for candidate in candidates:
        entity_id = candidate.get("id")
        if not entity_id:
            continue

        entity_params = {
            "action": "wbgetentities",
            "format": "json",
            "languages": "en",
            "props": "claims",
            "ids": entity_id,
        }
        entity_url = f"{WIKIDATA_API}?{urllib.parse.urlencode(entity_params)}"

        try:
            entity_payload = fetch_json(entity_url)
        except Exception:
            continue

        entity = entity_payload.get("entities", {}).get(entity_id, {})
        claims = entity.get("claims", {})
        nationality_claims = claims.get("P27", [])

        for claim in nationality_claims:
            mainsnak = claim.get("mainsnak", {})
            datavalue = mainsnak.get("datavalue", {})
            value = datavalue.get("value", {})
            country_qid = value.get("id")
            if not country_qid:
                continue

            country_params = {
                "action": "wbgetentities",
                "format": "json",
                "languages": "en",
                "props": "labels",
                "ids": country_qid,
            }
            country_url = f"{WIKIDATA_API}?{urllib.parse.urlencode(country_params)}"

            try:
                country_payload = fetch_json(country_url)
            except Exception:
                continue

            country_entity = country_payload.get("entities", {}).get(country_qid, {})
            label = country_entity.get("labels", {}).get("en", {}).get("value")
            code = normalize_country_code(label)
            if code and label:
                return code, label

    return None, None


def guess_country_from_author(author, custom_map):
    """Guess country from author name using known authors map + custom overrides."""
    # Check custom map first
    if author in custom_map:
        return custom_map[author]

    # Check built-in map
    if author in AUTHOR_COUNTRY_MAP:
        return AUTHOR_COUNTRY_MAP[author]

    # If not found, return None
    return None


def parse_goodreads_csv(
    csv_path,
    custom_map,
    country_cache,
    title_overrides,
    title_cache,
    posts_lookup,
    instagram_posts,
    cover_lookup,
    bookclub_lookup,
    cover_cache,
):
    """Parse Goodreads export CSV and return book entries."""
    books = []
    unresolved = []
    lookup_budget = {"remaining": 220}
    cover_lookup_budget = {"remaining": 30}

    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)

            for row_idx, row in enumerate(reader, start=2):
                # Only include books marked as "read"
                exclusive_shelf = row.get("Exclusive Shelf", "").strip()
                if exclusive_shelf != "read":
                    continue

                title = row.get("Title", "").strip()
                author = row.get("Author", "").strip()
                date_read = row.get("Date Read", "").strip()
                year_published = row.get("Year Published", "").strip()
                book_id_source = row.get("Book Id", "").strip()
                rating = row.get("My Rating", "").strip()
                my_review = row.get("My Review", "").strip()

                if not title or not author:
                    print(f"Warning: Row {row_idx} missing title or author. Skipping.")
                    continue

                country_code = None
                country_label = None

                # 1) Manual/custom map
                mapped = guess_country_from_author(author, custom_map)
                if mapped:
                    country_code = normalize_country_code(mapped)
                    country_label = country_label_from_code(country_code)

                # 2) Persistent cache
                if not country_code and author in country_cache:
                    cached_code = normalize_country_code(country_cache[author])
                    if cached_code:
                        country_code = cached_code
                        country_label = country_label_from_code(country_code)

                # 3) Internet lookup (Wikidata)
                if not country_code:
                    net_code, net_label = lookup_country_via_wikidata(author)
                    if net_code:
                        country_code = net_code
                        country_label = net_label or country_label_from_code(net_code)
                        country_cache[author] = net_code

                # 4) Final fallback
                if not country_code:
                    country_code = "unknown"
                    country_label = "Unknown"
                    unresolved.append(f"  - {title} by {author}")

                # Create book entry
                book_id = f"{slugify(title)}-{slugify(author.split()[-1])}"
                if book_id_source.isdigit():
                    goodreads_url = f"https://www.goodreads.com/book/show/{book_id_source}"
                else:
                    goodreads_url = None

                local_rating = int(rating) if rating.isdigit() and int(rating) > 0 else None
                english_title = resolve_english_title(
                    row,
                    title,
                    author,
                    title_overrides,
                    title_cache,
                    lookup_budget,
                )

                curated_post, matched_instagram_post = resolve_post_metadata(
                    title,
                    english_title,
                    author,
                    posts_lookup,
                    instagram_posts,
                )

                default_note_en = (
                    "I read this book before I started writing regular notes on books. "
                    "I am keeping this entry as a memory marker in my reading journey."
                )
                default_note_tr = (
                    "Bu kitabi kitaplar uzerine duzenli notlar tutmaya baslamadan once okumustum. "
                    "Bu kaydi okuma yolculugumun bir hafiza izi olarak koruyorum."
                )

                curated_note_en = (curated_post or {}).get("note_en", "").strip()
                curated_note_tr = (curated_post or {}).get("note_tr", "").strip()
                curated_note_generic = (curated_post or {}).get("note", "").strip()
                if curated_note_generic and not curated_note_en and not curated_note_tr:
                    if is_likely_turkish_text(curated_note_generic):
                        curated_note_tr = curated_note_generic
                    else:
                        curated_note_en = curated_note_generic

                review_text_en = curated_note_en
                review_text_tr = curated_note_tr

                if my_review:
                    if is_likely_turkish_text(my_review):
                        if not review_text_tr:
                            review_text_tr = my_review
                    else:
                        if not review_text_en:
                            review_text_en = my_review

                instagram_caption = (matched_instagram_post or {}).get("caption", "").strip()
                if instagram_caption:
                    if is_likely_turkish_text(instagram_caption):
                        if not review_text_tr:
                            review_text_tr = instagram_caption
                    else:
                        if not review_text_en:
                            review_text_en = instagram_caption

                if not review_text_en and not review_text_tr:
                    review_text_en = default_note_en
                    review_text_tr = default_note_tr

                review_text_en, review_text_tr = build_bilingual_note_pair(review_text_en, review_text_tr)

                curated_quotes_en = list((curated_post or {}).get("quotes_en", []) or [])
                curated_quotes_tr = list((curated_post or {}).get("quotes_tr", []) or [])
                curated_quotes_generic = list((curated_post or {}).get("quotes", []) or [])
                if curated_quotes_generic and not curated_quotes_en and not curated_quotes_tr:
                    inferred_en, inferred_tr = split_quotes_by_language(curated_quotes_generic)
                    curated_quotes_en = inferred_en
                    curated_quotes_tr = inferred_tr
                quote_candidates_en, quote_candidates_tr = build_bilingual_quote_pair(curated_quotes_en, curated_quotes_tr)

                place_date = (curated_post or {}).get("place_date", "")
                if place_date:
                    if review_text_en and review_text_en not in {EN_NOTE_TR_ONLY, TR_NOTE_EN_ONLY}:
                        review_text_en = f"{review_text_en}\n\n{place_date}"
                    if review_text_tr and review_text_tr not in {EN_NOTE_TR_ONLY, TR_NOTE_EN_ONLY}:
                        review_text_tr = f"{review_text_tr}\n\n{place_date}"

                tags_en = ["goodreads", "library"]
                author_tag = keywordify(author)
                country_tag = keywordify(country_label)
                year_tag = keywordify(year_published)
                for candidate in [author_tag, country_tag, year_tag]:
                    if candidate and candidate not in tags_en:
                        tags_en.append(candidate)

                for keyword in (curated_post or {}).get("keywords", []):
                    keyword_token = keywordify(keyword)
                    if keyword_token and keyword_token not in tags_en:
                        tags_en.append(keyword_token)

                if (curated_post or {}).get("add_to_bookclub") and "book-club" not in tags_en:
                    tags_en.append("book-club")

                essay_en = f"<h3>Reading Note</h3><p>{review_text_en}</p>"
                essay_tr = (
                    f"<h3>Okuma Notu</h3><p>{review_text_tr}</p>"
                )

                local_cover_url = resolve_local_cover(
                    [
                        title,
                        english_title,
                        (curated_post or {}).get("title_tr", ""),
                        (curated_post or {}).get("title_en", ""),
                    ],
                    cover_lookup,
                )

                bookclub_url = resolve_bookclub_url(
                    title,
                    english_title,
                    author,
                    (curated_post or {}).get("add_to_bookclub", False),
                    bookclub_lookup,
                )

                # Build entry
                entry = {
                    "id": book_id,
                    "type": "book",
                    "title": {"en": english_title, "tr": title},
                    "creator": author,
                    "country": country_code,
                    "countryLabel": {"en": country_label, "tr": country_label},
                    "year": year_published or "Unknown",
                    "readDate": date_read if date_read else None,
                    "coverUrl": local_cover_url or build_cover_url(row, cover_cache, cover_lookup_budget),
                    "goodreadsUrl": goodreads_url,
                    "bookClubUrl": bookclub_url,
                    "rating": f"{local_rating}/5" if local_rating else None,
                    "note": {
                        "en": review_text_en,
                        "tr": review_text_tr,
                    },
                    "essay": {
                        "en": essay_en,
                        "tr": essay_tr,
                    },
                    "quotes": {
                        "en": quote_candidates_en,
                        "tr": quote_candidates_tr,
                    },
                    "tags": {
                        "en": tags_en,
                        "tr": tags_en,
                    },
                }

                books.append(entry)

    except FileNotFoundError:
        print(f"Error: {csv_path} not found.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error parsing CSV: {e}", file=sys.stderr)
        sys.exit(1)

    if unresolved:
        print(f"\nCountry unresolved for {len(unresolved)} books:")
        for item in unresolved:
            print(item)
        print("\nTip: add overrides in book-country-map.json")

    return books


def generate_js_output(books, output_path):
    """Generate JavaScript module from book data."""
    js_content = f"""// Auto-generated from Goodreads export CSV
// Generated: {datetime.now().isoformat()}
// Edit goodreads_library_export.csv and re-run generate_book_dataset.py to update

const bookEntries = {json.dumps(books, ensure_ascii=False, indent=2)};
"""

    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(js_content)
        print(f"✓ Generated {output_path} with {len(books)} books")
    except Exception as e:
        print(f"Error writing output: {e}", file=sys.stderr)
        sys.exit(1)


def generate_example_country_map():
    """Generate an example book-country-map.json if it doesn't exist."""
    map_file = Path("book-country-map.json")
    if not map_file.exists():
        example = {
            "John Author": "unitedStates",
            "Jane Writer": "france",
            "#": "Copy this file and add your own author -> country mappings",
        }
        try:
            with open(map_file, "w", encoding="utf-8") as f:
                json.dump(example, f, indent=2)
            print(f"✓ Created example {map_file}")
        except Exception as e:
            print(f"Warning: Could not create {map_file}: {e}")


def main():
    """Main entry point."""
    csv_path = ROOT / "data" / "imports" / "goodreads" / "goodreads_library_export.csv"
    js_output = ROOT / "data" / "generated" / "auto" / "book-data.js"

    if not csv_path.exists():
        print(f"Error: {csv_path} not found.", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {csv_path}...")
    custom_map = load_custom_country_map()
    title_overrides = load_title_override_map()
    country_cache = load_country_cache()
    title_cache = load_title_cache()
    cover_cache = load_cover_cache()
    posts_rows = load_posts_excel_rows()
    posts_lookup = build_posts_excel_lookup(posts_rows)
    instagram_posts = parse_instagram_posts_html()
    cover_lookup = load_local_cover_lookup()
    bookclub_lookup = load_bookclub_lookup()

    print(
        f"Loaded metadata: {len(posts_rows)} Excel post rows, "
        f"{len(instagram_posts)} Instagram posts, "
        f"{len(cover_lookup)} local covers"
    )

    books = parse_goodreads_csv(
        csv_path,
        custom_map,
        country_cache,
        title_overrides,
        title_cache,
        posts_lookup,
        instagram_posts,
        cover_lookup,
        bookclub_lookup,
        cover_cache,
    )
    save_country_cache(country_cache)
    save_title_cache(title_cache)
    save_cover_cache(cover_cache)

    if not books:
        print(
            "\nError: No books found. Make sure goodreads_library_export.csv has books marked as 'read' "
            "in the Exclusive Shelf column.",
            file=sys.stderr,
        )
        generate_example_country_map()
        sys.exit(1)

    # Sort by read date (newest first)
    books_sorted = sorted(
        books,
        key=lambda b: (b.get("readDate") or "0000-00-00"),
        reverse=True,
    )

    generate_js_output(books_sorted, js_output)

    # Print summary
    print(f"\nSummary:")
    print(f"  Total books read: {len(books_sorted)}")
    countries = set(b["country"] for b in books_sorted)
    print(f"  Countries: {len(countries)}")
    print(
        f"\nNext time you add a new book to Goodreads:"
        f"\n  1. Mark it as 'read' in your Goodreads library"
        f"\n  2. Export your Goodreads library"
        f"\n  3. Run: python3 generate_book_dataset.py"
        f"\n\nIf a country looks wrong or unresolved:"
        f"\n  - Edit or create book-country-map.json"
        f"\n  - Add: \"Author Name\": \"countryCode\""
        f"\n  - Run the script again"
        f"\n\nIf an English title looks wrong:"
        f"\n  - Edit or create book-title-map.json"
        f"\n  - Add: \"GoodreadsBookId\": \"English Title\""
        f"\n  - Run the script again"
    )


if __name__ == "__main__":
    main()
